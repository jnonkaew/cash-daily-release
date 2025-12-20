import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
import os
import json

# ================== CONFIG ==================
CONFIG_FILE = "config.json"
SAFE_DEDUCT = 3000

def load_last_folder():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data.get("last_folder", "")
        except:
            return ""
    return ""

def save_last_folder(folder):
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump({"last_folder": folder}, f, ensure_ascii=False, indent=2)
    except:
        pass

# ================== GLOBAL ==================
loaded_file_path = None
file_date_str = ""  # เก็บวันที่จากไฟล์

def focus_next(event):
    event.widget.tk_focusNext().focus()
    return "break"

# ================== NUMBER FORMAT (CEILING) ==================
from decimal import Decimal, ROUND_CEILING
import re

def bind_number_format(entry):
    var = tk.StringVar()
    busy = False

    def validate(P):
        if P == "":
            return True

        # ❌ อนุญาตเฉพาะตัวเลข , และ .
        if not re.fullmatch(r"[0-9,\.]+", P):
            return False

        # ❌ ห้ามมี . มากกว่า 1 ตัว
        if P.count(".") > 1:
            return False

        # ❌ ถ้ามี . ต้องไม่ใช่ตัวแรก
        if P.startswith("."):
            return False

        # ❌ ตัวหน้าของ . ต้องเป็นตัวเลข หรือ ,
        if "." in P:
            idx = P.index(".")
            if P[idx - 1] not in "0123456789,":
                return False

        # ลบ , กับ . แล้วต้องเหลือเป็นตัวเลข
        if not P.replace(",", "").replace(".", "").isdigit():
            return False

        return True

    vcmd = (entry.register(validate), "%P")
    entry.config(
        validate="key",
        validatecommand=vcmd,
        textvariable=var
    )

    def format_value(event=None):
        nonlocal busy
        if busy:
            return

        value = var.get().replace(",", "")
        if value == "" or value == ".":
            return

        busy = True
        try:
            num = Decimal(value).quantize(
                Decimal("1"), rounding=ROUND_CEILING
            )
            var.set(f"{int(num):,}")
        except:
            pass
        busy = False

    entry.bind("<FocusOut>", format_value)
    entry.bind("<Return>", format_value)

    return var


def num(v):
    try:
        return int(
            Decimal(v.get().replace(",", "")).quantize(
                Decimal("1"), rounding=ROUND_CEILING
            )
        )
    except:
        return 0

# ================== DATE ==================
def format_date_th(d):
    months = ["ม.ค.","ก.พ.","มี.ค.","เม.ย.","พ.ค.","มิ.ย.",
              "ก.ค.","ส.ค.","ก.ย.","ต.ค.","พ.ย.","ธ.ค."]
    if isinstance(d, str):
        try:
            d = datetime.strptime(d, "%Y-%m-%d")
        except:
            return d
    return f"{d.day} {months[d.month-1]} {d.year+543}"

# ================== GUI ==================
root = tk.Tk()  # <<< ต้องสร้าง root ก่อน
root.title("ระบบบันทึกเงินสดร้านโชว์ห่วย V1.1")
win_width = 500
win_height = 660
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = (screen_width - win_width) // 2
y = (screen_height - win_height) // 2
root.geometry(f"{win_width}x{win_height}+{x}+{y}")
root.minsize(win_width, win_height)
root.rowconfigure(3, weight=1)
root.columnconfigure(0, weight=1)

# ================== OPEN FOLDER ==================
def open_loaded_folder(event=None):
    if not loaded_file_path:
        return
    folder = os.path.dirname(loaded_file_path)
    if os.path.exists(folder):
        os.startfile(folder)

# ================== FILE FUNCTIONS ==================
def choose_folder():
    folder = filedialog.askdirectory(initialdir=save_path.get() or load_last_folder())
    if folder:
        save_path.set(folder)
        save_last_folder(folder)

def get_unique_filename(folder, base):
    i = 0
    while True:
        name = f"{base}{'' if i == 0 else f'_{i:02d}'}.xlsx"
        path = os.path.join(folder, name)
        if not os.path.exists(path):
            return path
        i += 1

# ================== LOAD ==================
def load_excel():
    global loaded_file_path, file_date_str
    path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if not path:
        return
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    def setv(v, val):
        v.set(f"{int(val):,}" if val else "")

    setv(e_cash_start, ws["B3"].value)
    setv(e_iship, ws["B4"].value)
    setv(e_flashhome, ws["B5"].value)
    setv(e_showhuiay, ws["B6"].value)
    setv(e_ttb, ws["B7"].value)
    setv(e_gsb, ws["B8"].value)
    setv(e_cash_actual, ws["B11"].value)

    loaded_file_path = path

    file_date = ws["B1"].value
    if file_date:
        if isinstance(file_date, str):
            try:
                dt = datetime.strptime(file_date, "%Y-%m-%d")
                file_date_str = format_date_th(dt)
            except:
                file_date_str = file_date
        else:
            file_date_str = format_date_th(file_date)
    else:
        file_date_str = ""

    lbl_file.config(text=f"📄 ไฟล์ที่ดึงมา: {path}")
    save_path.set(os.path.dirname(os.path.dirname(path)))
    save_last_folder(save_path.get())
    btn_update.config(state="normal")
    update_summary()

def ensure_today_date():
    global file_date_str
    if not file_date_str:
        file_date_str = format_date_th(datetime.now())

# ================== SUMMARY ==================
def update_summary():
    ensure_today_date()
    summary.delete("1.0", tk.END)
    total_sales = num(e_iship) + num(e_flashhome) + num(e_showhuiay)
    total_transfer = num(e_ttb) + num(e_gsb)
    summary.insert(tk.END, "สรุปยอดเงินสดประจำวัน\n")
    summary.insert(tk.END, f"วันที่ {file_date_str or '-'}\n")
    summary.insert(tk.END, "*****************************\n")
    summary.insert(tk.END, f"ยอดเงินก่อนเริ่มงาน      {num(e_cash_start):,} บาท\n", "bigbold")
    summary.insert(tk.END, f"ยอดขายรวม   -----------  {total_sales:,} บาท\n")
    summary.insert(tk.END, f"   - ระบบ iShip          {num(e_iship):>11,}\n")
    summary.insert(tk.END, f"   - ระบบ FlashHome   {num(e_flashhome):>8,}\n")
    summary.insert(tk.END, f"   - ระบบ ShowHuaiy   {num(e_showhuiay):>9,}\n\n")
    summary.insert(tk.END, "-------------------------------\n\n")
    summary.insert(tk.END, f"เงินโอน   -----------   {total_transfer:,} บาท\n", "bigbold")
    summary.insert(tk.END, f"   - เงินโอนเข้า TTB      {num(e_ttb):>8,}\n")
    summary.insert(tk.END, f"   - เงินโอนเข้า GSB      {num(e_gsb):>8,} {'(รอโอน)' if gsb_pending.get() else '(โอนแล้ว)'}\n\n")
    summary.insert(tk.END, "-------------------------------\n")
    summary.insert(tk.END, f"เงินสดนับได้ล่าสุด        {num(e_cash_actual):,} บาท\n")
    summary.insert(tk.END, "-------------------------------\n")
    
    cash_safe = max(num(e_cash_actual) - SAFE_DEDUCT, 0)
    
    summary.insert(tk.END,f"เงินทอน                     {SAFE_DEDUCT:>8,} บาท\n","bigbold")
    summary.insert(tk.END,f"ยอดเงินใส่ตู้เซฟ           {cash_safe:>7,} บาท\n","bigbold")

def copy_summary():
    root.clipboard_clear()
    root.clipboard_append(summary.get("1.0", tk.END))
    messagebox.showinfo("คัดลอกแล้ว", "คัดลอกรายงานเรียบร้อย")

# ================== RESET ==================
def reset_form():
    global loaded_file_path, file_date_str

    # เคลียร์ค่าช่องตัวเลขทั้งหมด
    for v in (
        e_cash_start,
        e_iship,
        e_flashhome,
        e_showhuiay,
        e_ttb,
        e_gsb,
        e_cash_actual
    ):
        v.set("")

    # ค่าเริ่มต้น checkbox
    gsb_pending.set(True)

    # เคลียร์ summary
    summary.delete("1.0", tk.END)

    # รีเซ็ตสถานะไฟล์
    loaded_file_path = None
    file_date_str = ""

    # รีเซ็ตข้อความไฟล์
    lbl_file.config(text="📄 ไฟล์ที่ดึงมา: ยังไม่ได้ดึงไฟล์")

    # ปิดปุ่มแก้ไขไฟล์
    btn_update.config(state="disabled")

    # ❌ ไม่แตะ save_path (โฟลเดอร์ปลายทางยังอยู่)


# ================== SAVE ==================
def save_excel():
    if not save_path.get():
        messagebox.showerror("ผิดพลาด", "กรุณาเลือกโฟลเดอร์ปลายทาง")
        return
    try:
        dt = datetime.strptime(file_date_str, "%d %b %Y")
    except:
        dt = datetime.now()
    month_folder = os.path.join(save_path.get(), dt.strftime("%Y%m"))
    os.makedirs(month_folder, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 32
    ws["A1"] = "รายงานตรวจสอบเงินสดประจำวัน ที่"
    ws["B1"] = dt
    ws["B1"].number_format = '[$-th-TH]d mmm yyyy'
    ws["A1"].font = ws["B1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = ws["B1"].alignment = Alignment(vertical="center")
    ws["A3"], ws["B3"] = "ยอดเงินในเก๊ะก่อนเริ่มงาน", num(e_cash_start)
    ws["A4"], ws["B4"] = "ยอดขายระบบ iShip", num(e_iship)
    ws["A5"], ws["B5"] = "ยอดขายระบบ FlashHome", num(e_flashhome)
    ws["A6"], ws["B6"] = "ยอดขายระบบ ShowHuaiy", num(e_showhuiay)
    ws["A7"], ws["B7"] = "เงินโอนเข้า TTB", num(e_ttb)
    ws["A8"], ws["B8"] = "เงินโอนเข้า GSB", num(e_gsb)
    ws["A11"], ws["B11"] = "ยอดเงินนับได้จริงในเก๊ะ", num(e_cash_actual)
    
    cash_safe = max(num(e_cash_actual) - SAFE_DEDUCT, 0)
    
    ws["A12"], ws["B12"] = "ยอดเงินใส่ตู้เซฟ", "=B11-3000"
    
    file_name = get_unique_filename(month_folder, dt.strftime("%Y%m%d"))
    wb.save(file_name)
    messagebox.showinfo("สำเร็จ", f"บันทึกแล้ว\n{os.path.basename(file_name)}")

# ================== UPDATE ==================
def update_excel():
    if not loaded_file_path:
        return
    if not messagebox.askyesno("ยืนยัน", "ต้องการแก้ไขไฟล์เดิมหรือไม่?"):
        return
    wb = load_workbook(loaded_file_path)
    ws = wb.active
    ws["B3"] = num(e_cash_start)
    ws["B4"] = num(e_iship)
    ws["B5"] = num(e_flashhome)
    ws["B6"] = num(e_showhuiay)
    ws["B7"] = num(e_ttb)
    ws["B8"] = num(e_gsb)
    ws["B11"] = num(e_cash_actual)
    wb.save(loaded_file_path)
    messagebox.showinfo("สำเร็จ", "แก้ไขไฟล์เดิมเรียบร้อย")

# ================== GUI WIDGETS ==================
lbl_file = tk.Label(root, text="📄 ไฟล์ที่ดึงมา: ยังไม่ได้ดึงไฟล์", fg="blue", cursor="hand2", anchor="w", wraplength=540)
lbl_file.grid(row=0, column=0, sticky="ew", padx=10, pady=4)
lbl_file.bind("<Button-1>", open_loaded_folder)
lbl_file.bind("<Enter>", lambda e: lbl_file.config(fg="#0b5ed7"))
lbl_file.bind("<Leave>", lambda e: lbl_file.config(fg="blue"))

frame = tk.Frame(root)
frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=4)
frame.columnconfigure(1, weight=1)

def create_row(label, r):
    tk.Label(frame, text=label).grid(row=r, column=0, sticky="w", padx=5, pady=4)
    e = tk.Entry(frame, width=30, justify="right")
    e.grid(row=r, column=1, sticky="ew", padx=5, pady=4)
    v = bind_number_format(e)
    e.bind("<KeyRelease>", lambda e: update_summary())
    e.bind("<Return>", focus_next)


    return v

e_cash_start = create_row("ยอดเงินก่อนเริ่มงาน", 0)
frame.grid_slaves(row=0, column=1)[0].focus_set()
e_iship = create_row("ยอดขายระบบ iShip", 1)
e_flashhome = create_row("ยอดขายระบบ FlashHome", 2)
e_showhuiay = create_row("ยอดขายระบบ ShowHuaiy", 3)
e_ttb = create_row("เงินโอนเข้า TTB", 4)

tk.Label(frame, text="เงินโอนเข้า GSB").grid(row=5, column=0, sticky="w", padx=5, pady=4)
e_gsb_entry = tk.Entry(frame, width=30, justify="right")
e_gsb_entry.grid(row=5, column=1, sticky="ew", padx=5, pady=4)
e_gsb = bind_number_format(e_gsb_entry)
e_gsb_entry.bind("<KeyRelease>", lambda e: update_summary())
e_gsb_entry.bind("<Return>", focus_next)
e_cash_actual = create_row("เงินสดนับได้ล่าสุด", 6)

# ================== Check box รอโอน ==================
gsb_pending = tk.BooleanVar(value=True)
tk.Checkbutton(frame, text="รอผ้าไทยโอน", variable=gsb_pending, command=update_summary).grid(row=5, column=2, padx=5, pady=4)



save_path = tk.StringVar()
save_path.set(load_last_folder())
tk.Label(frame, text="โฟลเดอร์ปลายทาง").grid(row=7, column=0, sticky="w", padx=5, pady=4)
tk.Entry(frame, textvariable=save_path, width=30).grid(row=7, column=1, sticky="ew", padx=5, pady=4)
tk.Button(frame, text="เลือกโฟลเดอร์", command=choose_folder)\
    .grid(row=7, column=2, padx=(2, 5), pady=4, sticky="w")


btns = tk.Frame(root)
btns.grid(row=2, column=0, pady=4)
tk.Button(btns, text="📂 ดึงข้อมูล", bg="#6c757d", fg="white", command=load_excel).grid(row=0, column=0, padx=3)
tk.Button(btns, text="💾 บันทึกใหม่", bg="#2e86de", fg="white", command=save_excel).grid(row=0, column=1, padx=3)
btn_update = tk.Button(btns, text="✏️ แก้ไขไฟล์", bg="#f39c12", fg="white", command=update_excel, state="disabled")
btn_update.grid(row=0, column=2, padx=3)
tk.Button(btns, text="📋 คัดลอก", bg="#27ae60", fg="white", command=copy_summary).grid(row=0, column=3, padx=3)
tk.Button(btns, text="♻️ รีเซ็ต", bg="#c0392b", fg="white", command=reset_form).grid(row=0, column=4, padx=3)

sum_frame = tk.Frame(root)
sum_frame.grid(row=3, column=0, sticky="nsew", padx=10, pady=6)
sum_frame.rowconfigure(0, weight=1)
sum_frame.columnconfigure(0, weight=1)

scroll = tk.Scrollbar(sum_frame)
scroll.grid(row=0, column=1, sticky="ns")

summary = tk.Text(sum_frame, font=("Consolas", 11), yscrollcommand=scroll.set)
summary.grid(row=0, column=0, sticky="nsew")
summary.tag_config("bigbold", font=("Consolas", 12, "bold"))
scroll.config(command=summary.yview)

root.mainloop()
